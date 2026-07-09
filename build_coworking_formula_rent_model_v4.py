from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUT = "C:/tmp/coworking_general_model_v4.zip"
FINAL = "coworking_general_model_v4.xlsx"
COMP = "경쟁사/경쟁사 및 부동산 매물 분석_.xlsx"

ROOMS = [("1인실", 8), ("2인실", 4), ("4인실", 14)]
TIERS = [
    ("A. 저임대 권역", 30000, 300000, 870000, 1340000, "지방 광역시/수도권 외곽 우량 매물", "적극 검토"),
    ("B. 표준 권역", 40000, 300000, 870000, 1340000, "분당·수도권 일반 업무권역", "표준모델"),
    ("C. 핵심 권역", 50000, 330000, 930000, 1430000, "서울 외곽 업무권역/광역시 핵심권역", "선별 검토"),
    ("D. 상급 권역", 60000, 350000, 1000000, 1550000, "판교·여의도 등 수요 강한 권역", "예외 검토"),
    ("E. 고가 권역", 70000, 400000, 1100000, 1700000, "강남·선릉 등 고임대 권역", "원칙 제외"),
]
LISTINGS = [
    ("분당", "분당.png", "일반상가", 393, 433, "전용 393㎡, 월세 433만원"),
    ("분당", "분당100평 수정.png", "일반상가", 347, 440, "전용 347㎡, 월세 440만원"),
    ("분당", "분당100평 수정.png", "일반상가", 398, 435, "전용 398㎡, 월세 435만원"),
    ("분당", "분당100평 수정.png", "일반상가", 393, 460, "전용 393㎡, 월세 460만원"),
    ("분당", "분당100평.png", "일반상가", 322, 530, "전용 322㎡, 월세 530만원"),
    ("분당", "분당 큰평수.png", "일반상가", 470, 560, "전용 470㎡, 월세 560만원"),
    ("분당", "분당 큰평수.png", "일반상가", 480, 560, "전용 480㎡, 월세 560만원"),
    ("여의도", "여의도.png", "일반상가", 297, 750, "전용 297㎡, 월세 750만원"),
    ("여의도", "여의도.png", "일반상가", 297, 700, "전용 297㎡, 월세 700만원"),
    ("여의도", "여의도.png", "일반상가", 297, 730, "전용 297㎡, 월세 730만원"),
    ("여의도", "여의도1.png", "사무실", 304, 620, "전용 304㎡, 월세 620만원"),
    ("판교", "판교.png", "복합상가", 344, 650, "전용 344㎡, 월세 650만원"),
    ("판교", "판교1.png", "일반상가", 344, 600, "전용 344㎡, 월세 600만원"),
    ("선릉", "선릉.png", "일반상가", 296, 750, "전용 296㎡, 월세 750만원"),
    ("강남", "강남빌딩2.png", "대형사무실", 89, 320, "100평급 표본 아님"),
    ("강남", "강남빌딩1.png", "중소형사무실", 144, 800, "고가 표본"),
    ("강남", "강남빌딩1.png", "대형사무실", 297, 1700, "고가 표본"),
]
CAPEX_ITEMS = [
    ("인테리어/방음 공사", 130000000, "소형 독립실 칸막이, 방음, 바닥/천장/도장"),
    ("가구/집기", 25000000, "책상, 의자, 수납, 회의실 집기"),
    ("네트워크/OA/보안", 15000000, "인터넷, Wi-Fi, 복합기, CCTV, 출입통제"),
    ("설계/사인/소방 등", 12500000, "설계, 사인, 소방/전기 보완"),
    ("예비비/원복충당", 20000000, "예상 외 공사 및 원상복구 대비"),
]

# 실매물 확정 3개 규모 모델 — standard_100.mjs / pangyo_120.mjs / pangyo_150.mjs 실행결과와 동일 (2026-07-09 검증)
# (규모, 전용평, 실수, 좌석수, 4인실수, 2인실수, 1인실수, 4인실가_만원, 2인실가_만원, 1인실가_만원, 월세_만원, 인건비_만원, 파트타임인건비_만원, 출처)
SIZE_MODELS = [
    ("100평형", 105, 26, 72, 14, 4, 8, 134, 87, 30, 440, 350, 0, "standard_100.mjs (실매물 전용105평·월세440만)"),
    ("120평형", 120, 32, 86, 17, 3, 12, 123, 62, 41, 432, 350, 0, "pangyo_120.mjs (실매물 전용120평·월세432만)"),
    ("150평형", 150, 52, 106, 17, 3, 32, 123, 62, 41, 560, 350, 120, "pangyo_150.mjs (실매물 전용150평·월세560만, 파트타임 인건비 별도)"),
]


def style(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border
            if c.row == 1:
                c.font = Font(bold=True, color="FFFFFF", size=12)
                c.fill = PatternFill("solid", fgColor="1F4E78")
            elif c.row == 2 or c.value in ("항목", "티어", "지역", "시나리오", "객실", "구분"):
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="D9EAF7")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 22
    ws.freeze_panes = "A3"


def mark(ws, cells):
    fill = PatternFill("solid", fgColor="FFF2CC")
    for addr in cells:
        ws[addr].fill = fill


def add_size_model(wb):
    ws = wb.create_sheet("00_규모모델", 0)
    ws.append(["규모별(100·120·150평) 실매물 확정 모델 — 월 손익 한눈에 보기"])
    ws.append(["항목"] + [m[0] for m in SIZE_MODELS])
    cols = ["B", "C", "D"]

    def add_row(label, values):
        ws.append([label] + values)

    add_row("실수(호실)", [m[2] for m in SIZE_MODELS])          # row3
    add_row("좌석수", [m[3] for m in SIZE_MODELS])              # row4
    add_row("4인실 수", [m[4] for m in SIZE_MODELS])            # row5
    add_row("2인실 수", [m[5] for m in SIZE_MODELS])            # row6
    add_row("1인실 수", [m[6] for m in SIZE_MODELS])            # row7
    add_row("4인실가(만원)", [m[7] for m in SIZE_MODELS])       # row8
    add_row("2인실가(만원)", [m[8] for m in SIZE_MODELS])       # row9
    add_row("1인실가(만원)", [m[9] for m in SIZE_MODELS])       # row10
    add_row("전용평수", [m[1] for m in SIZE_MODELS])            # row11
    add_row("월세(만원)", [m[10] for m in SIZE_MODELS])         # row12
    add_row("전용평당월세(만원/평)", [f"={c}12/{c}11" for c in cols])  # row13
    add_row("만실매출(만원/월)", [f"={c}5*{c}8+{c}6*{c}9+{c}7*{c}10" for c in cols])  # row14
    add_row("관리비(만원, 1.5만/평)", [f"=ROUND({c}11*1.5,0)" for c in cols])  # row15
    add_row("전기(만원, 0.6만/평)", [f"=ROUND({c}11*0.6,0)" for c in cols])  # row16
    add_row("청소(만원)", [f"=ROUND({c}11*0.4+{c}4*1,0)" for c in cols])  # row17
    add_row("운영잡비(만원)", [32, 32, 32])                     # row18
    add_row("인건비(만원)", [m[11] for m in SIZE_MODELS])       # row19
    add_row("파트타임 인건비(만원)", [m[12] for m in SIZE_MODELS])  # row20
    add_row("CAPEX(만원)", [f"=ROUND(({c}11*163+{c}3*7+{c}4*23.65+286)*1.05,0)" for c in cols])  # row21
    add_row("CAPEX 월상각(만원)", [f"=ROUND({c}21/60,0)" for c in cols])  # row22
    add_row("월 고정비 합계(만원)", [f"={c}12+{c}15+{c}16+{c}17+{c}18+{c}19+{c}20+{c}22" for c in cols])  # row23
    add_row("변동비율(마케팅+PG 등)", [0.115, 0.115, 0.115])    # row24
    add_row("BEP(손익분기 가동률)", [f"={c}23/({c}14*(1-{c}24))" for c in cols])  # row25
    ws.append([])                                               # row26
    ws.append(["가동률 70% 시나리오"])                          # row27
    add_row("70% 월매출(만원)", [f"={c}14*0.7" for c in cols])  # row28
    add_row("(–) 변동비(만원)", [f"={c}28*{c}24" for c in cols])  # row29
    add_row("(–) 월 고정비(만원)", [f"={c}23" for c in cols])   # row30
    add_row("= 70% 월손익(만원)", [f"={c}28-{c}29-{c}30" for c in cols])  # row31
    ws.append([])                                               # row32
    ws.append(["가동률 90% 시나리오"])                          # row33
    add_row("90% 월매출(만원)", [f"={c}14*0.9" for c in cols])  # row34
    add_row("(–) 변동비(만원)", [f"={c}34*{c}24" for c in cols])  # row35
    add_row("(–) 월 고정비(만원)", [f"={c}23" for c in cols])   # row36
    add_row("= 90% 월손익(만원)", [f"={c}34-{c}35-{c}36" for c in cols])  # row37
    ws.append([])                                               # row38
    ws.append(["출처"] + [m[13] for m in SIZE_MODELS])          # row39
    style(ws)


def add_summary(wb):
    ws = wb.create_sheet("01_Summary")
    ws.append(["요약"])
    ws.append(["항목", "내용", "비고"])
    ws.append(["모델", "100·120·150평 임차형 소형특화 공유오피스 (1·2·4인실 중심, 지정석 제외)", "실매물 확정 3개 모델 — 00_규모모델 시트 참조"])
    ws.append([])
    ws.append(["규모별 월 손익 — 한눈에 보기 (00_규모모델 원자료)"])
    ws.append(["구분", "100평형", "120평형", "150평형"])
    ws.append(["70% 가동 월매출(만원)", "='00_규모모델'!B28", "='00_규모모델'!C28", "='00_규모모델'!D28"])
    ws.append(["(–) 변동비(만원)", "='00_규모모델'!B29", "='00_규모모델'!C29", "='00_규모모델'!D29"])
    ws.append(["(–) 월 고정비(만원)", "='00_규모모델'!B30", "='00_규모모델'!C30", "='00_규모모델'!D30"])
    ws.append(["= 70% 월손익(만원)", "='00_규모모델'!B31", "='00_규모모델'!C31", "='00_규모모델'!D31"])
    ws.append([])
    ws.append(["90% 가동 월매출(만원)", "='00_규모모델'!B34", "='00_규모모델'!C34", "='00_규모모델'!D34"])
    ws.append(["(–) 변동비(만원)", "='00_규모모델'!B35", "='00_규모모델'!C35", "='00_규모모델'!D35"])
    ws.append(["(–) 월 고정비(만원)", "='00_규모모델'!B36", "='00_규모모델'!C36", "='00_규모모델'!D36"])
    ws.append(["= 90% 월손익(만원)", "='00_규모모델'!B37", "='00_규모모델'!C37", "='00_규모모델'!D37"])
    ws.append([])
    ws.append(["참고: BEP(손익분기 가동률)", "='00_규모모델'!B25", "='00_규모모델'!C25", "='00_규모모델'!D25"])  # row17
    ws.append([])
    ws.append(["참고 — 미확보 매물 지역 스크리닝용 임대료 티어 시나리오 (면적 100평 가정, 위 3개 실매물 확정모델과는 별개 도구)"])
    ws.append(["티어", "전용평당 월세", "권장 1인실", "권장 2인실", "권장 4인실", "적용 권역", "만실매출", "월 고정비", "BEP", "70% 월손익", "90% 월손익", "판단"])
    for src in range(31, 36):
        ws.append([f"='02_RentModel'!A{src}", f"='02_RentModel'!B{src}", f"='02_RentModel'!C{src}", f"='02_RentModel'!D{src}", f"='02_RentModel'!E{src}", f"='02_RentModel'!F{src}", f"='02_RentModel'!H{src}", f"='02_RentModel'!I{src}", f"='02_RentModel'!J{src}", f"='02_RentModel'!K{src}", f"='02_RentModel'!L{src}", f"='02_RentModel'!M{src}"])
    # 위 티어 데이터는 row21~25, BEP는 열I
    style(ws)


def add_rent_model(wb):
    ws = wb.create_sheet("02_RentModel")
    ws.append(["임대료 표준모델"])
    ws.append(["지역", "파일", "유형", "전용㎡", "월세_만원", "전용평", "전용평당월세_만원", "비고"])
    for item in LISTINGS:
        row = ws.max_row + 1
        ws.append([item[0], item[1], item[2], item[3], item[4], f"=D{row}/3.3058", f"=E{row}/F{row}", item[5]])
    mark(ws, [f"D{i}" for i in range(3, 20)] + [f"E{i}" for i in range(3, 20)])
    ws.append([])
    ws.append(["지역별 임대료 요약"])
    ws.append(["지역", "표본수", "최저_만원/평", "평균_만원/평", "최고_만원/평"])
    for region in ["분당", "여의도", "판교", "선릉", "강남"]:
        row = ws.max_row + 1
        ws.append([region, f'=COUNTIF($A$3:$A$19,A{row})', f'=MINIFS($G$3:$G$19,$A$3:$A$19,A{row})', f'=AVERAGEIF($A$3:$A$19,A{row},$G$3:$G$19)', f'=MAXIFS($G$3:$G$19,$A$3:$A$19,A{row})'])
    ws.append([])
    ws.append(["전국 확장용 임대료 티어"])
    ws.append(["티어", "전용평당 월세", "권장 1인실", "권장 2인실", "권장 4인실", "적용 권역", "전용평수", "만실매출", "월 고정비", "BEP", "70% 월손익", "90% 월손익", "판단"])
    for tier, rent, p1, p2, p4, desc, decision in TIERS:
        row = ws.max_row + 1
        ws.append([tier, rent, p1, p2, p4, desc, "='04_CAPEX'!B3", f"='03_Revenue'!B3*C{row}+'03_Revenue'!B4*D{row}+'03_Revenue'!B5*E{row}", f"=G{row}*B{row}+'04_CAPEX'!B17", f"=I{row}/(H{row}*(1-'04_CAPEX'!B18))", f"=H{row}*0.7*(1-'04_CAPEX'!B18)-I{row}", f"=H{row}*0.9*(1-'04_CAPEX'!B18)-I{row}", decision])
    mark(ws, [f"B{i}" for i in range(31, 36)] + [f"C{i}" for i in range(31, 36)] + [f"D{i}" for i in range(31, 36)] + [f"E{i}" for i in range(31, 36)])
    style(ws)


def add_revenue(wb):
    ws = wb.create_sheet("03_Revenue")
    ws.append(["객실 수량 및 티어별 매출"])
    ws.append(["객실", "실수", "비고"])
    for name, count in ROOMS:
        ws.append([name, count, "수량 입력값"])
    mark(ws, ["B3", "B4", "B5"])
    ws.append([])
    ws.append(["가격/매출 기준", "02_RentModel의 임대료 티어별 권장 가격을 사용", "지역/임대료 티어에 따라 만실매출 변동"])
    ws.append([])
    ws.append(["티어", "1인실가", "2인실가", "4인실가", "만실매출", "계산식"])
    for src in range(31, 36):
        ws.append([f"='02_RentModel'!A{src}", f"='02_RentModel'!C{src}", f"='02_RentModel'!D{src}", f"='02_RentModel'!E{src}", f"=B3*B{ws.max_row+1}+B4*C{ws.max_row+1}+B5*D{ws.max_row+1}", "1인실수*1인실가 + 2인실수*2인실가 + 4인실수*4인실가"])
    style(ws)


def add_capex(wb):
    ws = wb.create_sheet("04_CAPEX")
    ws.append(["CAPEX 및 고정비 상세"])
    ws.append(["항목", "금액/값", "단위", "산정 근거"])
    ws.append(["전용평수", 100, "평", "표준모델 면적"])
    ws.append(["상각기간", 60, "개월", "CAPEX 월상각 기준"])
    start = ws.max_row + 1
    for item in CAPEX_ITEMS:
        ws.append([item[0], item[1], "원", item[2]])
    ws.append(["CAPEX 합계", f"=SUM(B{start}:B{start+len(CAPEX_ITEMS)-1})", "원", "세부항목 합계"])
    ws.append(["CAPEX 월상각", f"=B{start+len(CAPEX_ITEMS)}/B4", "원", "CAPEX 합계 / 상각기간"])
    ws.append(["관리비 단가", 30000, "원/평", "월 관리비 가정"])
    ws.append(["전기·냉난방 단가", 8000, "원/평", "월 유틸리티 가정"])
    ws.append(["관리비 월액", "=B3*B12", "원", "전용평수 x 관리비 단가"])
    ws.append(["전기·냉난방 월액", "=B3*B13", "원", "전용평수 x 유틸리티 단가"])
    ws.append(["공통 운영비", 3700000, "원", "운영 280만원 + 보험·렌탈·원복충당 90만원"])
    ws.append(["임대료 제외 고정비", "=B11+B14+B15+B16", "원", "월상각+관리비+유틸리티+공통운영비"])
    ws.append(["변동비·마케팅률", 0.115, "%", "마케팅 7% + PG·플랫폼·환불 4.5%"])
    mark(ws, ["B3", "B4", "B12", "B13", "B16", "B18"])
    for r in range(start, start + len(CAPEX_ITEMS)):
        mark(ws, [f"B{r}"])
    style(ws)


def add_pl(wb):
    ws = wb.create_sheet("05_PL")
    ws.append(["입주율/공실률별 손익"])
    ws.append(["티어", "입주율", "공실률", "만실매출", "월매출", "월 고정비", "변동비", "월영업이익", "연영업이익"])
    out = 3
    for src in range(31, 36):
        for occ in [0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
            ws.append([f"='02_RentModel'!A{src}", occ, f"=1-B{out}", f"='02_RentModel'!H{src}", f"=D{out}*B{out}", f"='02_RentModel'!I{src}", f"=E{out}*'04_CAPEX'!B18", f"=E{out}-F{out}-G{out}", f"=H{out}*12"])
            out += 1
    style(ws)


def add_sales(wb):
    ws = wb.create_sheet("06_Sales")
    ws.append(["권역 세일즈 인건비 배부"])
    ws.append(["기준 지점", "시나리오", "권역 세일즈 월 총비용", "담당 지점 수", "지점당 배부비", "BEP", "90% 월영업이익", "비고"])
    scenarios = [("초기 본사 겸임", 0, 0, "초기 1~2개 지점"), ("3개 지점당 1명", 4500000, 3, "영업 강화"), ("4개 지점당 1명", 4500000, 4, "기준"), ("5개 지점당 1명", 4500000, 5, "비용 효율")]
    for name, cost, branches, note in scenarios:
        row = ws.max_row + 1
        alloc = "0" if branches == 0 else f"=C{row}/D{row}"
        ws.append(["B. 표준 권역 100평 지점", name, cost, branches, alloc, f"=('02_RentModel'!I32+E{row})/('02_RentModel'!H32*(1-'04_CAPEX'!B18))", f"='02_RentModel'!H32*0.9*(1-'04_CAPEX'!B18)-('02_RentModel'!I32+E{row})", note])
    mark(ws, ["C4", "C5", "C6", "D4", "D5", "D6"])
    style(ws)


def add_competitor(wb):
    ws = wb.create_sheet("07_Competitor")
    ws.append(["경쟁사 분석 정리"])
    ws.append(["지역", "경쟁사", "객실/상품", "경쟁사 월가", "경쟁사 인당", "SOS 비교상품", "SOS 월가", "SOS 인당", "공실/공급 메모", "판단"])
    rows = [
        ("판교", "패스트파이브", "4인실", 3090000, "=D3/4", "상급권역 4인실", "='02_RentModel'!E34", "=G3/4", "1·2·4인실 공실 없음 메모", '=IF(H3<E3,"가격 우위","가격 열위")'),
        ("판교", "패스트파이브", "7인실", 4060000, "=D4/7", "상급권역 4인실", "='02_RentModel'!E34", "=G4/4", "대형 체인 기준가격", '=IF(H4<E4,"가격 우위","가격 열위")'),
        ("분당", "스파크플러스", "5인실 파격가", 1850000, 370000, "표준권역 4인실", "='02_RentModel'!E32", "=G5/4", "분당 1·2·4 공실 없음 메모", '=IF(H5<E5,"가격 우위","가격 열위")'),
        ("야탑", "슈가맨워크", "4인실", 850000, 212500, "표준권역 4인실", "='02_RentModel'!E32", "=G6/4", "2인실 1개, 4인실 1개 공실 메모", '=IF(H6<E6,"가격 우위","가격 열위")'),
        ("분당", "더분당", "4인실", 800000, 200000, "표준권역 4인실", "='02_RentModel'!E32", "=G7/4", "총 36호실, 4인실 총 4실 메모", '=IF(H7<E7,"가격 우위","가격 열위")'),
    ]
    for row in rows:
        ws.append(list(row))
    style(ws)


def add_source_and_check(wb):
    ws = wb.create_sheet("08_RE_Source")
    ws.append(["부동산 매물 원자료"])
    ws.append(["지역", "파일", "유형", "전용㎡", "월세_만원", "전용평", "전용평당월세_만원", "비고"])
    for item in LISTINGS:
        row = ws.max_row + 1
        ws.append([item[0], item[1], item[2], item[3], item[4], f"=D{row}/3.3058", f"=E{row}/F{row}", item[5]])
    style(ws)
    ws = wb.create_sheet("09_Check")
    ws.append(["수동 검산표"])
    ws.append(["티어", "1인실가", "2인실가", "4인실가", "만실매출", "월고정비", "BEP", "70% 월손익", "90% 월손익"])
    area, mgmt, util, capex_month, common, var = 100, 30000, 8000, 202500000 / 60, 3700000, 0.115
    for tier, rent, p1, p2, p4, desc, decision in TIERS:
        revenue = 8 * p1 + 4 * p2 + 14 * p4
        fixed = area * rent + mgmt * area + util * area + capex_month + common
        ws.append([tier, p1, p2, p4, revenue, fixed, fixed / (revenue * (1 - var)), revenue * 0.7 * (1 - var) - fixed, revenue * 0.9 * (1 - var) - fixed])
    style(ws)


wb = Workbook()
wb.remove(wb.active)
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
add_size_model(wb)
add_summary(wb)
add_rent_model(wb)
add_revenue(wb)
add_capex(wb)
add_pl(wb)
add_sales(wb)
add_competitor(wb)
add_source_and_check(wb)

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, (int, float)) or (isinstance(c.value, str) and c.value.startswith("=")):
                c.number_format = "#,##0"
            if ws.title in ("02_RentModel", "05_PL", "06_Sales", "09_Check") and c.column_letter in ("B", "C", "G", "J"):
                if "BEP" in str(ws.cell(2, c.column).value) or c.coordinate in ("J31", "J32", "J33", "J34", "J35"):
                    c.number_format = "0.0%"
            if ws.title == "00_규모모델" and c.row in (24, 25) and c.column_letter in ("B", "C", "D"):
                c.number_format = "0.0%"
            if ws.title == "01_Summary":
                if c.row == 17 and c.column_letter in ("B", "C", "D"):
                    c.number_format = "0.0%"
                if 21 <= c.row <= 25 and c.column_letter == "I":
                    c.number_format = "0.0%"

wb.save(OUT)
print(OUT)


