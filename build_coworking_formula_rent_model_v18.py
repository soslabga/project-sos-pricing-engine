from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

OUT = "C:/tmp/coworking_general_model_v18.zip"
FINAL = "coworking_general_model_v18.xlsx"
LAYOUT_IMAGES = [
    ("100평형", "배치도_100평.png"),
    ("120평형", "배치도_120평.png"),
    ("150평형", "배치도_150평.png"),
]
COMP = "경쟁사/경쟁사 및 부동산 매물 분석__.xlsx"

ROOMS = [("1인실", 8), ("2인실", 4), ("4인실", 14)]
TIERS = [
    ("A. 저임대 권역", 30000, 300000, 870000, 1340000, "지방 광역시/수도권 외곽 우량 매물", "적극 검토"),
    ("B. 표준 권역", 40000, 300000, 870000, 1340000, "분당·수도권 일반 업무권역", "표준모델"),
    ("C. 핵심 권역", 50000, 330000, 930000, 1430000, "서울 외곽 업무권역/광역시 핵심권역", "선별 검토"),
    ("D. 상급 권역", 60000, 350000, 1000000, 1550000, "판교·여의도 등 수요 강한 권역", "예외 검토"),
    ("E. 고가 권역", 70000, 400000, 1100000, 1700000, "강남·선릉 등 고임대 권역", "원칙 제외"),
]
LISTINGS = [
    ("분당", "분당.png", "일반상가", 393, 433, "전용 393㎡, 월세 433만원"),
    ("분당", "분당100평 수정.png", "일반상가", 347, 440, "전용 347㎡, 월세 440만원"),
    ("분당", "분당100평 수정.png", "일반상가", 398, 435, "전용 398㎡, 월세 435만원"),
    ("분당", "분당100평 수정.png", "일반상가", 393, 460, "전용 393㎡, 월세 460만원"),
    ("분당", "분당100평.png", "일반상가", 322, 530, "전용 322㎡, 월세 530만원"),
    ("분당", "분당 큰평수.png", "일반상가", 470, 560, "전용 470㎡, 월세 560만원"),
    ("분당", "분당 큰평수.png", "일반상가", 480, 560, "전용 480㎡, 월세 560만원"),
    ("여의도", "여의도.png", "일반상가", 297, 750, "전용 297㎡, 월세 750만원"),
    ("여의도", "여의도.png", "일반상가", 297, 700, "전용 297㎡, 월세 700만원"),
    ("여의도", "여의도.png", "일반상가", 297, 730, "전용 297㎡, 월세 730만원"),
    ("여의도", "여의도1.png", "사무실", 304, 620, "전용 304㎡, 월세 620만원"),
    ("판교", "판교.png", "복합상가", 344, 650, "전용 344㎡, 월세 650만원"),
    ("판교", "판교1.png", "일반상가", 344, 600, "전용 344㎡, 월세 600만원"),
    ("선릉", "선릉.png", "일반상가", 296, 750, "전용 296㎡, 월세 750만원"),
    ("강남", "강남빌딩2.png", "대형사무실", 89, 320, "100평급 표본 아님"),
    ("강남", "강남빌딩1.png", "중소형사무실", 144, 800, "고가 표본"),
    ("강남", "강남빌딩1.png", "대형사무실", 297, 1700, "고가 표본"),
]
CAPEX_ITEMS = [
    ("인테리어/방음 공사", 130000000, "소형 독립실 칸막이, 방음, 바닥/천장/도장"),
    ("가구/집기", 25000000, "책상, 의자, 수납, 회의실 집기"),
    ("네트워크/OA/보안", 15000000, "인터넷, Wi-Fi, 복합기, CCTV, 출입통제"),
    ("설계/사인/소방 등", 12500000, "설계, 사인, 소방/전기 보완"),
    ("예비비/원복충당", 20000000, "예상 외 공사 및 원상복구 대비"),
]

# 실매물 확정 3개 규모 모델 — standard_100.mjs / pangyo_120.mjs / pangyo_150.mjs 실행결과와 동일 (2026-07-09 검증)
# 방구성/전용평수/CAPEX는 각 실매물 배치도 기준 고정값. 가격(4/2/1인실가·월세)은 00_규모모델(분당3종)에서 지역 티어 드롭다운으로 재계산됨
# (규모, 전용평, 실수, 좌석수, 4인실수, 2인실수, 1인실수, 4인실가_만원, 2인실가_만원, 1인실가_만원, 실제계약월세_만원, 인건비_만원, 파트타임인건비_만원, 출처)
SIZE_MODELS = [
    ("100평형(분당)", 105, 26, 72, 14, 4, 8, 134, 87, 30, 440, 350, 0, "08_부동산원자료 시트 원본 대조(2026.7.9)"),
    ("120평형(분당)", 120, 32, 86, 17, 3, 12, 123, 62, 41, 435, 350, 0, "08_부동산원자료 시트 원본 대조(2026.7.9)"),
    ("150평형(분당)", 145, 52, 106, 17, 3, 32, 123, 62, 41, 560, 350, 120, "08_부동산원자료 시트 원본 대조(2026.7.9)"),
]


def style(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border
            if c.row == 1:
                c.font = Font(bold=True, color="FFFFFF", size=12)
                c.fill = PatternFill("solid", fgColor="1F4E78")
            elif c.row == 2 or c.value in ("항목", "티어", "지역", "시나리오", "객실", "구분"):
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="D9EAF7")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 22
    ws.freeze_panes = "A3"


def mark(ws, cells):
    fill = PatternFill("solid", fgColor="FFF2CC")
    for addr in cells:
        ws[addr].fill = fill


def add_layout(wb):
    # 100/120/150평형 실제 배치도(SVG→PNG 렌더, 지역명·가격 텍스트 제거된 중립본). 방구성은 00_규모모델(분당3종) SIZE_MODELS와 정확히 일치(2026-07-10 재검증).
    ws = wb.create_sheet("00_배치도", 1)
    ws["A1"] = "규모별(100·120·150평) 배치도 — 실제 배치도 기준(방구성은 00_규모모델(분당3종)과 동일)"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.column_dimensions["A"].width = 22
    target_width_px = 1000
    row_px = 20
    cur_row = 3
    for label, filename in LAYOUT_IMAGES:
        with PILImage.open(filename) as im:
            w, h = im.size
        scale = target_width_px / w
        disp_w, disp_h = target_width_px, round(h * scale)
        cell = ws.cell(cur_row, 1, f"{label} 배치도")
        cell.font = Font(bold=True, size=13, color="1F4E78")
        img = XLImage(filename)
        img.width, img.height = disp_w, disp_h
        ws.add_image(img, f"A{cur_row + 1}")
        cur_row += 1 + (disp_h // row_px) + 3


def add_common_assumptions(wb):
    # 공통 가정(정책 상수)은 별도 탭으로 분리 — 00_규모모델(분당3종) 맨 앞은 큰 그림(구성→매출→비용→손익)만 보이게
    ws = wb.create_sheet("00_공통가정", 2)
    ws.append(["공통 가정 — 3개 규모 모델(100·120·150평)에 동일 적용되는 정책 상수. 00_규모모델(분당3종) 시트 수식이 이 값들을 참조함"])  # row1
    ws.append(["항목", "값", "근거"])  # row2
    ws.append(["관리비 단가(만원/평)", 1.5, "출처 확인 불가 — 설계 가정치, 실사 시 재확인 필요"])  # row3
    ws.append(["전기 단가(만원/평)", 0.6, "출처 확인 불가 — 설계 가정치, 실사 시 재확인 필요"])  # row4
    ws.append(["청소 단가(만원/평)", 0.4, "청소용역비산출표 기준"])  # row5
    ws.append(["청소 단가(만원/좌석)", 1, "청소용역비산출표 기준"])  # row6
    ws.append(["운영잡비(만원, 정액)", 32, "인터넷4+복합기10+정수기3+CCTV보안12+화재보험3"])  # row7
    ws.append(["기본 인건비(만원, 현장담당 1인)", 350, "유사직무 연봉3600~3800만 기본급+법정부담 약20%"])  # row8
    ws.append(["변동비율(마케팅+PG 등)", 0.115, "마케팅 7% + PG·플랫폼·환불 4.5%"])  # row9
    ws.append(["초기투자비 상각기간(개월)", 60, ""])  # row10
    ws.append(["가동률 시나리오 A", 0.7, "표준 손익 확인용"])  # row11
    ws.append(["가동률 시나리오 B", 0.9, "안정화 이후 손익 확인용"])  # row12
    ws.append(["가동률 시나리오 C", 0.85, "회수기간 계산 전용 가동률. 90%(시나리오B, 안정화 이후 이론적 상한)를 쓰면 회수기간이 낙관적으로 짧게 나오고, 70%(시나리오A, 오픈초기 저가동 구간까지 포함한 평시 손익용)를 쓰면 이미 보수적으로 잡힌 가동률에 회수기간까지 또 보수화되어 과도하게 길게 나옴. '오픈 후 정상궤도 진입 이후 지속가능한 가동률'을 별도 분리"])  # row13
    ws.append(["운전자금(만원)", 500, "설계 가정치"])  # row14
    style(ws)
    mark(ws, [f"B{r}" for r in range(3, 15)])


def add_size_model(wb):
    ws = wb.create_sheet("00_규모모델(분당3종)", 0)
    cols = ["B", "C", "D"]
    CA = "'00_공통가정'!"

    def add_row(label, values):
        ws.append([label] + values)

    ws.append(["규모별(100·120·150평) 손익 모델 — 지역 티어 하나만 고르면 전체 재계산 (방구성·초기투자비는 각 실매물 배치도 고정, 가격만 티어 반영)"])  # row1
    ws.append(["지역 티어 선택", "B. 표준 권역", "", ""])  # row2 — 유일한 드롭다운, 아래 3개 평형 전체에 동시 적용
    ws.append(["항목"] + [m[0] for m in SIZE_MODELS])  # row3
    add_row("4인실 수", [m[4] for m in SIZE_MODELS])            # row4
    add_row("2인실 수", [m[5] for m in SIZE_MODELS])            # row5
    add_row("1인실 수", [m[6] for m in SIZE_MODELS])            # row6
    add_row("실수(호실)", [f"={c}4+{c}5+{c}6" for c in cols])           # row7
    add_row("좌석수", [f"={c}4*4+{c}5*2+{c}6*1" for c in cols])         # row8
    add_row("전용평수", [m[1] for m in SIZE_MODELS])            # row9
    add_row("전용평당월세(만원/평, 티어반영)", [f"=INDEX('02_임대료모델'!B31:B35,MATCH($B$2,'02_임대료모델'!A31:A35,0))/10000" for c in cols])  # row10
    add_row("월세(만원, 티어반영)", [f"={c}9*{c}10" for c in cols])  # row11
    add_row("4인실가(만원, 티어반영)", [f"=INDEX('02_임대료모델'!E31:E35,MATCH($B$2,'02_임대료모델'!A31:A35,0))/10000" for c in cols])  # row12
    add_row("2인실가(만원, 티어반영)", [f"=INDEX('02_임대료모델'!D31:D35,MATCH($B$2,'02_임대료모델'!A31:A35,0))/10000" for c in cols])  # row13
    add_row("1인실가(만원, 티어반영)", [f"=INDEX('02_임대료모델'!C31:C35,MATCH($B$2,'02_임대료모델'!A31:A35,0))/10000" for c in cols])  # row14
    add_row("만실매출(만원/월)", [f"={c}4*{c}12+{c}5*{c}13+{c}6*{c}14" for c in cols])  # row15
    add_row("관리비(만원)", [f"=ROUND({c}9*{CA}$B$3,0)" for c in cols])  # row16
    add_row("전기(만원)", [f"=ROUND({c}9*{CA}$B$4,0)" for c in cols])  # row17
    add_row("청소(만원)", [f"=ROUND({c}9*{CA}$B$5+{c}8*{CA}$B$6,0)" for c in cols])  # row18
    add_row("운영잡비(만원)", [f"={CA}$B$7" for c in cols])          # row19
    add_row("인건비(만원)", [f"={CA}$B$8" for c in cols])            # row20
    add_row("파트타임 인건비(만원)", [m[12] for m in SIZE_MODELS])  # row21 — 150평만 파트타임 추가(모델 고유 인력계획, 공통가정 아님)
    add_row("초기투자비(만원)", [f"=ROUND(({c}9*163+{c}7*7+{c}8*23.65+286)*1.05,0)" for c in cols])  # row22 — 세부 항목별 내역은 standard_100/pangyo_120/pangyo_150.mjs 주석 참조
    add_row("초기투자비 월상각(만원)", [f"=ROUND({c}22/{CA}$B$10,0)" for c in cols])  # row23
    add_row("월 고정비 합계(만원)", [f"={c}11+{c}16+{c}17+{c}18+{c}19+{c}20+{c}21+{c}23" for c in cols])  # row24
    add_row("변동비율(마케팅+PG 등)", [f"={CA}$B$9" for c in cols])  # row25
    add_row("BEP(손익분기 가동률)", [f"={c}24/({c}15*(1-{c}25))" for c in cols])  # row26
    ws.append([])  # row27
    ws.append([f'="가동률 "&TEXT({CA}$B$11,"0%")&" 시나리오"'])  # row28
    add_row(f'="가동률 "&TEXT({CA}$B$11,"0%")&" 월매출(만원)"', [f"={c}15*{CA}$B$11" for c in cols])  # row29
    add_row("(–) 변동비(만원)", [f"={c}29*{c}25" for c in cols])  # row30
    add_row("(–) 월 고정비(만원)", [f"={c}24" for c in cols])   # row31
    add_row(f'="= 가동률 "&TEXT({CA}$B$11,"0%")&" 월손익(만원)"', [f"={c}29-{c}30-{c}31" for c in cols])  # row32
    ws.append([])  # row33
    ws.append([f'="가동률 "&TEXT({CA}$B$12,"0%")&" 시나리오"'])  # row34
    add_row(f'="가동률 "&TEXT({CA}$B$12,"0%")&" 월매출(만원)"', [f"={c}15*{CA}$B$12" for c in cols])  # row35
    add_row("(–) 변동비(만원)", [f"={c}35*{c}25" for c in cols])  # row36
    add_row("(–) 월 고정비(만원)", [f"={c}24" for c in cols])   # row37
    add_row(f'="= 가동률 "&TEXT({CA}$B$12,"0%")&" 월손익(만원)"', [f"={c}35-{c}36-{c}37" for c in cols])  # row38
    ws.append([])  # row39
    add_row("참고: 실제 계약 월세(만원, 고정)", [m[10] for m in SIZE_MODELS])  # row40 — 지역이 이미 정해진 실계약 임대료. 위 티어반영 월세(row11)와는 별개 참고치
    add_row("보증금(만원, 고정)", [7500, 7500, 10000])  # row41 — 반환의무 부채(월세와 별개). 원본 출처 확인 불가한 설계 가정치(비고 참조)
    add_row("모델 성격(비고)", ["전국 확장 표준모델(분당 기준) — 월세는 실매물 검증, 보증금·룸믹스·CAPEX 단가는 설계 가정치"] * 3)  # row42
    ws.append(["출처"] + [m[13] for m in SIZE_MODELS])          # row43
    ws.append(["※ 방구성(4/2/1인실 수)·전용평수·초기투자비는 각 평형의 실제 배치도 기준 고정값이며, 지역 티어(B2)를 바꾸면 전용평당월세·객실가·월세(row10~14)만 재계산됩니다. 지역 티어(A~E) 자체는 미확보 매물 지역 참고용 가정. 임대차 계약기간·렌트프리 조건은 확인 불가 — 실사 재확인 필요."])  # row44

    ws["E3"] = "비고"
    notes = {
        4: "각 평형의 실제 도면상 방 구성 — 지역 선택과 무관하게 고정된 값",
        5: "각 평형의 실제 도면상 방 구성 — 지역 선택과 무관하게 고정된 값",
        6: "각 평형의 실제 도면상 방 구성 — 지역 선택과 무관하게 고정된 값",
        7: "전체 호실 수(4인실+2인실+1인실 합계)",
        8: "전체 좌석 수(4인실 4석·2인실 2석·1인실 1석으로 환산)",
        9: "실제 계약 기준 전용면적 — 지역 선택과 무관하게 고정된 값",
        10: "선택한 지역의 시세 임대료(평당) — 지역을 바꾸면 이 값이 바뀜",
        11: "전용평수 × 지역 시세 임대료로 추정한 월세. 실제 계약월세와는 다를 수 있음(비교치는 아래 '참고: 실제 계약 월세' 참조)",
        12: "선택한 지역의 권장 4인실 요금 — 지역별로 다르고, 평형(100/120/150평)과는 무관하게 3개 열 모두 동일",
        13: "선택한 지역의 권장 2인실 요금 — 3개 열 모두 동일",
        14: "선택한 지역의 권장 1인실 요금 — 3개 열 모두 동일",
        15: "방 구성 × 객실 요금으로 계산한 만실 기준 월매출",
        16: "관리비 = 전용평수 × 평당 관리비 단가(3개 평형 공통 정책값)",
        17: "전기료 = 전용평수 × 평당 전기 단가(공통 가정)",
        18: "청소비 = 전용평수·좌석수 기준 공통 단가 적용",
        19: "운영잡비(인터넷·복합기·정수기·보안·화재보험 등) 고정액 — 3개 평형 동일",
        20: "현장 담당 인건비 1인분 — 3개 평형 동일",
        21: "파트타임 인건비 — 150평형만 인력 추가 운영, 100·120평형은 없음",
        22: "CAPEX 산식(평수×163+실수×7+좌석수×23.65+286)의 단가 출처는 확인 불가 — 설계 가정치",
        23: "초기투자비를 60개월(5년)로 나눈 월 상각비",
        24: "월세+관리비+전기+청소+운영잡비+인건비+파트타임+초기투자비상각을 더한 월 총 고정비",
        25: "매출 대비 변동비율(마케팅+결제수수료 등, 공통 가정)",
        26: "손익분기 가동률 — 실제 가동률이 이 수치를 넘으면 흑자",
        28: "가동률 시나리오 A(기본 70%) 제목",
        29: "가동률 A 적용 시 월매출",
        30: "가동률 A 매출에 대한 변동비",
        31: "월 총 고정비(위와 동일)",
        32: "가동률 A 시나리오의 월 손익",
        34: "가동률 시나리오 B(기본 90%) 제목",
        35: "가동률 B 적용 시 월매출",
        36: "가동률 B 매출에 대한 변동비",
        37: "월 총 고정비(위와 동일)",
        38: "가동률 B 시나리오의 월 손익",
        40: "실제로 계약된 월세(고정값) — 위 '티어반영 월세'는 지역 선택에 따른 추정치라 실제 계약월세와 비교해보는 참고용",
        41: "업로드된 파일 중 원본 출처(mjs 파일) 확인 불가 — 설계 가정치, 실사 재확인 필요",
        42: "3개 평형 모두 분당 기준 전국 확장 표준모델. 월세는 08_부동산원자료 실매물로 검증됐으나 보증금·룸믹스·CAPEX 단가는 설계 가정치임",
        43: "각 평형의 방 구성·면적·초기투자비 산출 근거 자료",
    }
    for r, note in notes.items():
        ws[f"E{r}"] = note

    style(ws)
    for col in ("A", "B", "C", "D", "E"):
        ws[f"{col}3"].font = Font(bold=True)
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A4"
    mark(ws, [f"{c}{r}" for c in cols for r in (4, 5, 6, 9, 21, 40, 41, 42)])
    mark(ws, ["B2"])

    dv_tier = DataValidation(type="list", formula1="='02_임대료모델'!$A$31:$A$35", allow_blank=False)
    dv_tier.prompt = "지역 티어를 선택하세요"
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws["B2"])


def add_summary(wb):
    ws = wb.create_sheet("01_요약")
    ws.append(["요약"])
    ws.append(["항목", "내용", "비고"])
    ws.append(["모델", "100·120·150평 임차형 소형특화 공유오피스 (1·2·4인실 중심, 지정석 제외)", "실매물 확정 3개 모델 — 00_규모모델(분당3종) 시트 참조"])
    ws.append([])
    ws.append(["규모별 월 손익 — 한눈에 보기 (00_규모모델(분당3종) 원자료, 수식 연동, 지역 티어는 00_규모모델(분당3종)!B2 선택값 기준)"])
    ws.append(["구분"] + [m[0] for m in SIZE_MODELS])
    ws.append(["='00_규모모델(분당3종)'!A29", "='00_규모모델(분당3종)'!B29", "='00_규모모델(분당3종)'!C29", "='00_규모모델(분당3종)'!D29"])
    ws.append(["='00_규모모델(분당3종)'!A30", "='00_규모모델(분당3종)'!B30", "='00_규모모델(분당3종)'!C30", "='00_규모모델(분당3종)'!D30"])
    ws.append(["='00_규모모델(분당3종)'!A31", "='00_규모모델(분당3종)'!B31", "='00_규모모델(분당3종)'!C31", "='00_규모모델(분당3종)'!D31"])
    ws.append(["='00_규모모델(분당3종)'!A32", "='00_규모모델(분당3종)'!B32", "='00_규모모델(분당3종)'!C32", "='00_규모모델(분당3종)'!D32"])
    ws.append([])
    ws.append(["='00_규모모델(분당3종)'!A35", "='00_규모모델(분당3종)'!B35", "='00_규모모델(분당3종)'!C35", "='00_규모모델(분당3종)'!D35"])
    ws.append(["='00_규모모델(분당3종)'!A36", "='00_규모모델(분당3종)'!B36", "='00_규모모델(분당3종)'!C36", "='00_규모모델(분당3종)'!D36"])
    ws.append(["='00_규모모델(분당3종)'!A37", "='00_규모모델(분당3종)'!B37", "='00_규모모델(분당3종)'!C37", "='00_규모모델(분당3종)'!D37"])
    ws.append(["='00_규모모델(분당3종)'!A38", "='00_규모모델(분당3종)'!B38", "='00_규모모델(분당3종)'!C38", "='00_규모모델(분당3종)'!D38"])
    ws.append([])
    ws.append(["참고: BEP(손익분기 가동률)", "='00_규모모델(분당3종)'!B26", "='00_규모모델(분당3종)'!C26", "='00_규모모델(분당3종)'!D26"])  # row17
    ws.append([])
    ws.append(["참고 — 미확보 매물 지역 스크리닝용 임대료 티어 시나리오 (면적 100평 가정, 위 3개 실매물 확정모델과는 별개 도구)"])
    ws.append(["티어", "전용평당 월세", "권장 1인실", "권장 2인실", "권장 4인실", "적용 권역", "만실매출", "월 고정비", "BEP", "70% 월손익", "90% 월손익", "판단"])
    for src in range(31, 36):
        ws.append([f"='02_임대료모델'!A{src}", f"='02_임대료모델'!B{src}", f"='02_임대료모델'!C{src}", f"='02_임대료모델'!D{src}", f"='02_임대료모델'!E{src}", f"='02_임대료모델'!F{src}", f"='02_임대료모델'!H{src}", f"='02_임대료모델'!I{src}", f"='02_임대료모델'!J{src}", f"='02_임대료모델'!K{src}", f"='02_임대료모델'!L{src}", f"='02_임대료모델'!M{src}"])
    # 위 티어 데이터는 row21~25, BEP는 열I
    style(ws)


def add_rent_model(wb):
    ws = wb.create_sheet("02_임대료모델")
    ws.append(["임대료 표준모델"])
    ws.append(["지역", "파일", "유형", "전용㎡", "월세_만원", "전용평", "전용평당월세_만원", "비고"])
    for item in LISTINGS:
        row = ws.max_row + 1
        ws.append([item[0], item[1], item[2], item[3], item[4], f"=D{row}/3.3058", f"=E{row}/F{row}", item[5]])
    mark(ws, [f"D{i}" for i in range(3, 20)] + [f"E{i}" for i in range(3, 20)])
    ws.append([])
    ws.append(["지역별 임대료 요약"])
    ws.append(["지역", "표본수", "최저_만원/평", "평균_만원/평", "최고_만원/평"])
    for region in ["분당", "여의도", "판교", "선릉", "강남"]:
        row = ws.max_row + 1
        ws.append([region, f'=COUNTIF($A$3:$A$19,A{row})', f'=_xlfn.MINIFS($G$3:$G$19,$A$3:$A$19,A{row})', f'=AVERAGEIF($A$3:$A$19,A{row},$G$3:$G$19)', f'=_xlfn.MAXIFS($G$3:$G$19,$A$3:$A$19,A{row})'])
    ws.append([])
    ws.append(["전국 확장용 임대료 티어"])
    ws.append(["티어", "전용평당 월세", "권장 1인실", "권장 2인실", "권장 4인실", "적용 권역", "전용평수", "만실매출", "월 고정비", "BEP", "70% 월손익", "90% 월손익", "판단"])
    for tier, rent, p1, p2, p4, desc, decision in TIERS:
        row = ws.max_row + 1
        ws.append([tier, rent, p1, p2, p4, desc, "='04_초기투자비(구버전_미사용)'!B3", f"='03_매출'!B3*C{row}+'03_매출'!B4*D{row}+'03_매출'!B5*E{row}", f"=G{row}*B{row}+'04_초기투자비(구버전_미사용)'!B17", f"=I{row}/(H{row}*(1-'04_초기투자비(구버전_미사용)'!B18))", f"=H{row}*0.7*(1-'04_초기투자비(구버전_미사용)'!B18)-I{row}", f"=H{row}*0.9*(1-'04_초기투자비(구버전_미사용)'!B18)-I{row}", decision])
    mark(ws, [f"B{i}" for i in range(31, 36)] + [f"C{i}" for i in range(31, 36)] + [f"D{i}" for i in range(31, 36)] + [f"E{i}" for i in range(31, 36)])
    style(ws)


def add_revenue(wb):
    ws = wb.create_sheet("03_매출")
    ws.append(["객실 수량 및 티어별 매출"])
    ws.append(["객실", "실수", "비고"])
    for name, count in ROOMS:
        ws.append([name, count, "수량 입력값"])
    mark(ws, ["B3", "B4", "B5"])
    ws.append([])
    ws.append(["가격/매출 기준", "02_임대료모델의 임대료 티어별 권장 가격을 사용", "지역/임대료 티어에 따라 만실매출 변동"])
    ws.append([])
    ws.append(["티어", "1인실가", "2인실가", "4인실가", "만실매출", "계산식"])
    for src in range(31, 36):
        ws.append([f"='02_임대료모델'!A{src}", f"='02_임대료모델'!C{src}", f"='02_임대료모델'!D{src}", f"='02_임대료모델'!E{src}", f"=B3*B{ws.max_row+1}+B4*C{ws.max_row+1}+B5*D{ws.max_row+1}", "1인실수*1인실가 + 2인실수*2인실가 + 4인실수*4인실가"])
    style(ws)


def add_capex(wb):
    ws = wb.create_sheet("04_초기투자비(구버전_미사용)")
    ws.append(["본 시트는 00_규모모델(분당3종) 도입 이전 구버전이며, 관리비 단가(30,000원/평)가 00_공통가정(15,000원/평)과 다름. 현재 손익·투자 계산에는 사용되지 않으며 00_규모모델(분당3종)·10_투자회수 시트가 최신 기준임."])
    ws.append(["항목", "금액/값", "단위", "산정 근거"])
    ws.append(["전용평수", 100, "평", "표준모델 면적"])
    ws.append(["상각기간", 60, "개월", "초기투자비 월상각 기준"])
    start = ws.max_row + 1
    for item in CAPEX_ITEMS:
        ws.append([item[0], item[1], "원", item[2]])
    ws.append(["초기투자비 합계", f"=SUM(B{start}:B{start+len(CAPEX_ITEMS)-1})", "원", "세부항목 합계"])
    ws.append(["초기투자비 월상각", f"=B{start+len(CAPEX_ITEMS)}/B4", "원", "초기투자비 합계 / 상각기간"])
    ws.append(["관리비 단가", 30000, "원/평", "월 관리비 가정"])
    ws.append(["전기·냉난방 단가", 8000, "원/평", "월 유틸리티 가정"])
    ws.append(["관리비 월액", "=B3*B12", "원", "전용평수 x 관리비 단가"])
    ws.append(["전기·냉난방 월액", "=B3*B13", "원", "전용평수 x 유틸리티 단가"])
    ws.append(["공통 운영비", 3700000, "원", "운영 280만원 + 보험·렌탈·원복충당 90만원"])
    ws.append(["임대료 제외 고정비", "=B11+B14+B15+B16", "원", "월상각+관리비+유틸리티+공통운영비"])
    ws.append(["변동비·마케팅률", 0.115, "%", "마케팅 7% + PG·플랫폼·환불 4.5%"])
    mark(ws, ["B3", "B4", "B12", "B13", "B16", "B18"])
    for r in range(start, start + len(CAPEX_ITEMS)):
        mark(ws, [f"B{r}"])
    style(ws)


def add_pl(wb):
    ws = wb.create_sheet("05_손익")
    ws.append(["입주율/공실률별 손익"])
    ws.append(["티어", "입주율", "공실률", "만실매출", "월매출", "월 고정비", "변동비", "월영업이익", "연영업이익"])
    out = 3
    for src in range(31, 36):
        for occ in [0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
            ws.append([f"='02_임대료모델'!A{src}", occ, f"=1-B{out}", f"='02_임대료모델'!H{src}", f"=D{out}*B{out}", f"='02_임대료모델'!I{src}", f"=E{out}*'04_초기투자비(구버전_미사용)'!B18", f"=E{out}-F{out}-G{out}", f"=H{out}*12"])
            out += 1
    style(ws)


def add_sales(wb):
    ws = wb.create_sheet("06_영업인력")
    SM = "'00_규모모델(분당3종)'!"
    CA = "'00_공통가정'!"
    ws.append(["권역 세일즈 인건비 배부"])
    ws.append(["기준 지점", "시나리오", "권역 세일즈 월 총비용", "담당 지점 수", "지점당 배부비", "BEP", "90% 월영업이익", "비고"])
    scenarios = [("초기 본사 겸임", 0, 0, "초기 1~2개 지점"), ("3개 지점당 1명", 4500000, 3, "영업 강화"), ("4개 지점당 1명", 4500000, 4, "기준"), ("5개 지점당 1명", 4500000, 5, "비용 효율")]
    for name, cost, branches, note in scenarios:
        row = ws.max_row + 1
        alloc = "0" if branches == 0 else f"=C{row}/D{row}"
        ws.append([f"=\"100평형(분당) — \"&{SM}$B$2&\" 기준\"", name, cost, branches, alloc, f"=({SM}B24+E{row})/({SM}B15*(1-{CA}$B$9))", f"={SM}B15*0.9*(1-{CA}$B$9)-({SM}B24+E{row})", note])
    mark(ws, ["C4", "C5", "C6", "D4", "D5", "D6"])
    style(ws)


def add_competitor(wb):
    ws = wb.create_sheet("07_경쟁사분석")
    ws.append(["경쟁사 분석 정리 — 호실별 공실 현황 (경쟁사 및 부동산 매물 분석__.xlsx 원자료 기준, 2026-07-09 정리)"])
    ws.append(["출처 등급: [실거래]=계약서/입주제안서로 확인 · [정가]=경쟁사 공식 정가 · [현장확인]=직접 방문·상담으로 확인 · 공란=가격 미확보. "
               "조사 기준일 2026-07-09. 가격은 별도 표기 없는 한 부가세 별도(VAT excl.). 패스트파이브 크레딧(무료 회의실)은 데이터 미확보로 비교에 미반영."])
    ws.append(["지역", "경쟁사", "호실유형", "경쟁사 월가(원)", "출처등급", "총호실수", "공실수", "공실률", "SOS 비교가(원)", "비고"])
    P1, P2, P4 = "='00_규모모델(분당3종)'!B14*10000", "='00_규모모델(분당3종)'!B13*10000", "='00_규모모델(분당3종)'!B12*10000"
    # (지역, 경쟁사, 호실유형, 경쟁사월가_원, 출처등급, 총호실수, 공실수, SOS비교가, 비고) — 공실률은 아래에서 총호실수·공실수 둘 다 있을 때만 수식으로 채움
    rows = [
        ("판교", "패스트파이브", "1인실", None, "", None, 0, P1, "공실 없음(정확한 총실수는 원자료 미확인 — 확인 필요)"),
        ("판교", "패스트파이브", "2인실", None, "", None, 0, P2, "공실 없음(정확한 총실수는 원자료 미확인 — 확인 필요)"),
        ("판교", "패스트파이브", "4인실", 3090000, "[정가]", None, 0, P4, "공실 없음. 1.3 가격비교 기준(-57%). 실거래 할인가 없음 — 정가로 취급"),
        ("판교", "패스트파이브", "7인실", 4060000, "[정가]", None, None, None, "5평(0.89평/인), 공실현황 미기재. 실거래 할인가 없음 — 정가로 취급"),
        ("판교", "패스트파이브", "10인실", 5590000, "[정가]", None, None, None, "7평, 공실현황 미기재. 실거래 할인가 없음 — 정가로 취급"),
        ("분당", "스파크플러스", "1인실", None, "", 2, 0, P1, "인당 33만원(파격가) 별도 표기됨"),
        ("분당", "스파크플러스", "2인실", None, "", 4, 0, P2, ""),
        ("분당", "스파크플러스", "4인실", None, "", 5, 0, P4, ""),
        ("분당", "스파크플러스", "5인실", 1850000, "[현장확인]", None, None, None, "파격가, 인당 37만원"),
        ("분당", "스파크플러스", "6인실", None, "", None, None, None, "3.7평(사이즈만 확인, 가격·공실 미기재)"),
        ("분당", "스파크플러스", "8인실", None, "", None, None, None, "4.61평(사이즈만 확인, 가격·공실 미기재)"),
        ("야탑", "슈가맨워크(2호점)", "1인실", None, "", 2, None, P1, "공실현황 미기재"),
        ("야탑", "슈가맨워크(2호점)", "2인실", 550000, "[현장확인]", 2, 1, P2, ""),
        ("야탑", "슈가맨워크(2호점)", "3인실", None, "", 3, None, None, "공실현황 미기재"),
        ("야탑", "슈가맨워크(2호점)", "4인실", 850000, "[현장확인]", 3, 1, P4, ""),
        ("야탑", "이든비즈 야탑센터", "1인실", 300000, "[현장확인]", None, 1, P1, "총실수 미기재"),
        ("야탑", "이든비즈 야탑센터", "2인실", 500000, "[현장확인]", None, 1, P2, "총실수 미기재"),
        ("야탑", "이든비즈 야탑센터", "3인실", 600000, "[현장확인]", None, 1, None, "총실수 미기재"),
        ("야탑", "이든비즈 야탑센터", "4인실", None, "", None, 0, P4, "공실 없음"),
        ("분당", "리더스 분당(Regus)", "1인실", None, "", None, 3, P1, "인당 38만원. 1인실만 30% 할인 프로모션 중"),
        ("분당", "리더스 분당(Regus)", "2인실", None, "", None, 0, P2, "공실 없음"),
        ("분당", "리더스 분당(Regus)", "3인실", None, "", None, 1, None, ""),
        ("분당", "리더스 분당(Regus)", "4인실", None, "", None, 0, P4, "공실 없음"),
        ("분당", "더분당 공유오피스", "1인실", None, "", 23, 0, P1, "가격대 30~40만원, 공실 없음"),
        ("분당", "더분당 공유오피스", "2인실", 600000, "[현장확인]", 8, 1, P2, ""),
        ("분당", "더분당 공유오피스", "4인실", 800000, "[현장확인]", 4, 1, P4, ""),
        ("분당", "더분당 공유오피스", "5인실", None, "", 1, None, None, "공실현황 미기재. 총 36호실(23+8+4+1) 검증됨"),
        ("여의도", "스파크플러스", "1인실", None, "", None, 0, P1, "401~420호대 등 다수 호실 확인, 전부 공실 없음"),
        ("여의도", "스파크플러스", "2인실", 1300000, "[실거래]", None, 1, P2, "432호 실거래(정가180만→할인130만, VAT별도, 우리 입주 실계약서). 439호 공실(그 외 다수 호실 공실 없음). 실크기 2.959×1.953m"),
        ("여의도", "스파크플러스", "다인실(4·6인 등)", None, "", None, 3, None, "다인실 구간(401-420·501-525 등) 공실 3건. 5/8인 실거래(270만/335만)는 인실 구성이 달라 직접 비교 불가"),
    ]
    for region, comp, room_type, price, source_grade, total_rooms, vacant, sos_price, note in rows:
        r = ws.max_row + 1
        vacancy_rate = f"=G{r}/F{r}" if (total_rooms is not None and vacant is not None) else None
        ws.append([region, comp, room_type, price, source_grade, total_rooms, vacant, vacancy_rate, sos_price, note])
    style(ws)
    ws["A2"].font = Font(italic=True, color="647080")
    ws["A2"].fill = PatternFill(fill_type=None)
    for col in "ABCDEFGHIJ":
        ws[f"{col}3"].font = Font(bold=True)
        ws[f"{col}3"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws.row_dimensions[2].height = 42
    for r in range(4, 4 + len(rows)):
        ws[f"H{r}"].number_format = "0%"


def add_source_and_check(wb):
    ws = wb.create_sheet("08_부동산원자료")
    ws.append(["부동산 매물 원자료"])
    ws.append(["지역", "파일", "유형", "전용㎡", "월세_만원", "전용평", "전용평당월세_만원", "비고"])
    for item in LISTINGS:
        row = ws.max_row + 1
        ws.append([item[0], item[1], item[2], item[3], item[4], f"=D{row}/3.3058", f"=E{row}/F{row}", item[5]])
    style(ws)
    ws = wb.create_sheet("09_검산")
    ws.append(["수동 검산표"])
    ws.append(["티어", "1인실가", "2인실가", "4인실가", "만실매출", "월고정비", "BEP", "70% 월손익", "90% 월손익"])
    area, mgmt, util, capex_month, common, var = 100, 30000, 8000, 202500000 / 60, 3700000, 0.115
    for tier, rent, p1, p2, p4, desc, decision in TIERS:
        revenue = 8 * p1 + 4 * p2 + 14 * p4
        fixed = area * rent + mgmt * area + util * area + capex_month + common
        ws.append([tier, p1, p2, p4, revenue, fixed, fixed / (revenue * (1 - var)), revenue * 0.7 * (1 - var) - fixed, revenue * 0.9 * (1 - var) - fixed])
    style(ws)


def add_investment_recovery(wb):
    ws = wb.create_sheet("10_투자회수")
    cols = ["B", "C", "D"]
    CA = "'00_공통가정'!"
    SM = "'00_규모모델(분당3종)'!"

    def add_row(label, values):
        ws.append([label] + values)

    ws.append(["투자회수 분석 — 가동률 85% 시나리오 기준, 시설투자만 vs 총소요자금(시설투자+보증금+운전자금) 회수기간 비교"])  # row1
    ws.append(["항목"] + [m[0] for m in SIZE_MODELS])  # row2
    add_row("모델 성격", [f"={SM}{c}42" for c in cols])  # row3 — 100평형(분당)=전국 표준모델(대표값), 120·150평형(판교)=파일럿 참고
    add_row("만실매출(만원/월)", [f"={SM}{c}15" for c in cols])  # row4
    add_row("가동률 시나리오(85%)", [f"={CA}$B$13" for c in cols])  # row5
    add_row("월매출(만원, 85% 가동률)", [f"={c}4*{c}5" for c in cols])  # row6
    add_row("변동비율", [f"={SM}{c}25" for c in cols])  # row7
    add_row("(–) 변동비(만원)", [f"={c}6*{c}7" for c in cols])  # row8
    add_row("(–) 월 고정비(만원)", [f"={SM}{c}24" for c in cols])  # row9
    add_row("영업손익(만원, 85% 가동률)", [f"={c}6-{c}8-{c}9" for c in cols])  # row10
    add_row("(+) 초기투자비 월상각(만원)", [f"={SM}{c}23" for c in cols])  # row11
    add_row("월 현금흐름(영업손익+월상각비, 만원)", [f"={c}10+{c}11" for c in cols])  # row12
    ws.append([])  # row13
    add_row("초기투자비(CAPEX, 만원)", [f"={SM}{c}22" for c in cols])  # row14
    add_row("보증금(만원)", [f"={SM}{c}41" for c in cols])  # row15
    add_row("운전자금(만원)", [f"={CA}$B$14" for c in cols])  # row16
    add_row("총소요자금(만원, =CAPEX+보증금+운전자금)", [f"={c}14+{c}15+{c}16" for c in cols])  # row17
    ws.append([])  # row18
    add_row("회수기간 — 시설투자만 기준(개월)", [f"={c}14/{c}12" for c in cols])  # row19
    add_row("회수기간 — 총소요자금 기준(개월)", [f"={c}17/{c}12" for c in cols])  # row20
    add_row("회수기간 — 시설투자만 기준(년)", [f"={c}19/12" for c in cols])  # row21
    add_row("회수기간 — 총소요자금 기준(년)", [f"={c}20/12" for c in cols])  # row22
    ws.append([])  # row23
    ws.append(['="결론 — 분당 3개 규모(100/120/145평) 모두 전국 확장 표준모델로 사용 가능. 회수기간(총소요자금 기준): 100평형 "&TEXT(B22,"0.00")&"년, 120평형 "&TEXT(C22,"0.00")&"년, 150평형 "&TEXT(D22,"0.00")&"년"'])  # row24 — 총소요자금 기준 회수기간(년)은 B22/C22/D22 실시간 계산값 그대로 표시

    ws["E2"] = "비고"
    notes = {
        3: "00_규모모델(분당3종) 시트 42행 그대로 — 3개 평형 모두 분당 기준 전국 확장 표준모델",
        4: "00_규모모델(분당3종)의 만실 기준 월매출 그대로 가져옴(지역 티어 드롭다운 반영값)",
        5: "00_공통가정의 가동률 시나리오 C(85%) — 90%(이론적 상한)면 회수기간이 낙관적으로, 70%(평시용, 오픈초기 저가동 이미 반영)면 이중 보수화되어 과도하게 길게 나와 회수기간 계산 전용으로 별도 분리(00_공통가정 13행 근거 참조)",
        6: "만실매출 × 85% 가동률",
        7: "00_규모모델(분당3종)의 변동비율(마케팅+PG 등) 그대로 가져옴",
        8: "월매출 × 변동비율",
        9: "00_규모모델(분당3종)의 월 고정비 합계 그대로 가져옴",
        10: "월매출 - 변동비 - 월고정비",
        11: "초기투자비 월상각은 실제 현금 유출이 아니므로(감가상각) 영업손익에 다시 더해 현금흐름으로 환산",
        12: "영업손익 + 초기투자비 월상각(비현금 비용 환입) = 실제 월 현금창출액",
        14: "00_규모모델(분당3종)의 초기투자비(CAPEX) 그대로 가져옴",
        15: "반환의무 부채(월세와 별개). 원본 출처 확인 불가한 설계 가정치(00_규모모델(분당3종) 41행 비고 참조)",
        16: "설계 가정치",
        17: "초기투자비 + 보증금 + 운전자금 — 오픈까지 실제 필요한 총 조달액",
        19: "초기투자비만 회수하는 데 걸리는 개월 수(=CAPEX÷월 현금흐름)",
        20: "보증금·운전자금까지 포함한 총 조달액을 회수하는 데 걸리는 개월 수(=총소요자금÷월 현금흐름). 보증금은 반환의무 부채이지만 오픈 시점에 실제 지출되는 조달 필요액이므로 총소요자금 회수기간에는 포함",
        21: "위 회수기간(개월) ÷ 12",
        22: "위 회수기간(개월) ÷ 12",
    }
    for r, note in notes.items():
        ws[f"E{r}"] = note

    style(ws)
    for col in ("A", "B", "C", "D", "E"):
        ws[f"{col}2"].font = Font(bold=True)
        ws[f"{col}2"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws.column_dimensions["E"].width = 60
    ws.freeze_panes = "A3"


wb = Workbook()
wb.remove(wb.active)
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
add_size_model(wb)
add_layout(wb)
add_common_assumptions(wb)
add_summary(wb)
add_rent_model(wb)
add_revenue(wb)
add_capex(wb)
add_pl(wb)
add_sales(wb)
add_competitor(wb)
add_source_and_check(wb)
add_investment_recovery(wb)

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, (int, float)) or (isinstance(c.value, str) and c.value.startswith("=")):
                c.number_format = "#,##0"
            if ws.title in ("02_임대료모델", "05_손익", "06_영업인력", "09_검산") and c.column_letter in ("B", "C", "G", "J"):
                if "BEP" in str(ws.cell(2, c.column).value) or c.coordinate in ("J31", "J32", "J33", "J34", "J35"):
                    c.number_format = "0.0%"
            if ws.title == "06_영업인력" and c.column_letter == "F" and c.row >= 3:
                c.number_format = "0.0%"
            if ws.title == "00_공통가정" and c.row in (9, 11, 12, 13) and c.column_letter == "B":
                c.number_format = "0.0%"
            if ws.title == "00_규모모델(분당3종)" and c.row in (25, 26) and c.column_letter in ("B", "C", "D"):
                c.number_format = "0.0%"
            if ws.title == "01_요약":
                if c.row == 17 and c.column_letter in ("B", "C", "D"):
                    c.number_format = "0.0%"
                if 21 <= c.row <= 25 and c.column_letter == "I":
                    c.number_format = "0.0%"
            if ws.title == "07_경쟁사분석" and c.column_letter == "H" and c.row >= 4:
                c.number_format = "0%"
            if ws.title == "10_투자회수" and c.row in (5, 7) and c.column_letter in ("B", "C", "D"):
                c.number_format = "0.0%"

wb.save(OUT)
print(OUT)
